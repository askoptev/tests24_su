import openpyxl
import datetime
import csv
import json

"""
Модуль содержит функции для работы с файлами JSON XLSX CSV
The module contains functions for working with JSON XLSX CSV files
"""


def create_json(data_list=[]):
    """
    Create file JSON formats in to current directory
    """
    f_name = f'result_{datetime.date.today()}.json'
    try:
        with open(f_name, "a", encoding="utf-8") as file:
            json.dump(data_list, file, indent=4, ensure_ascii=False)
    except Exception as err:
        return err
    return f'create file: {f_name}'


def create_csv(data_list=[]):
    """
    Create file CSV formats in to current directory
    """
    f_name = f'result_{datetime.date.today()}.csv'
    try:
        with open(f_name, 'w', newline='') as csvfile:
            writer = csv.writer(csvfile, delimiter=';')
            writer.writerows(data_list)
    except Exception as err:
        return err
    return f'create file: {f_name}'


def create_xl(data_list=[], prefix='#'):
    """
    Create file Excel formats in to current directory
    """
    f_name = f'result_{datetime.date.today()}_{prefix}.xlsx'
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'main'
        row = 0
        col = 0
        for item in data_list:
            row += 1
            col = 0
            for value in item:
                col += 1
                cell = ws.cell(row=row, column=col)
                cell.value = value
        wb.save(f_name)
    except Exception as err:
        return err

    return f'create file: {f_name}'


def append_xl(data_list=[], prefix='#'):
    """
    Добавить данные в существующую книгу
    :param data_list:
    :param prefix:
    :return:
    """
    f_name = f'result_{datetime.date.today()}_{prefix}.xlsx'
    try:
        wb = openpyxl.load_workbook('test.xlsx')
        ws = wb.active
        for row in data_list:
            ws.append(row)
        wb.save(f_name)
    except Exception as err:
        return err
    return f'create file: {f_name}'


def append_csv(data_list=[]):
    f_name = f'result_{datetime.date.today()}.csv'
    try:
        with open(f_name, "a", encoding="utf-8") as csvfile:
            writer = csv.writer(csvfile, delimiter=';')
            writer.writerows(data_list)
    except Exception as err:
        return err
    return f'create file: {f_name}'

# screenshot = item.screenshot_as_png
# with open(f'browser_{inc}.png', 'wb') as f:
#     f.write(screenshot)
# inc += 1
